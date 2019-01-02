import os
import json
import openpyxl
import sys

def json_parser(file_path):
    fp = open(file_path, 'r', encoding='utf-8')
    #print("file-path = ",file_path)
    dict_info = json.load(fp)
    fp.close()
    return dict_info
##遍历指定文件夹下的所有文件，并找到对应json，一一对应写到case的Setting中
def AllAutoSetting(case_path,json_path):
    L_casename=[]
    L_jsonname=[]
    s_cutcase=[]
    s_cutjson=[]

   # def AutoSetting(case_path,json_path):
    L_casename = os.listdir(case_path)
    for casename in L_casename:
        s_cutcase=casename.split('_20')[0]
        #print(s_cutcase)
        L_jsonname= os.listdir(json_path)
        for jsonname in L_jsonname:
            s_cutjson=jsonname.split('.')[0]
            if s_cutjson in s_cutcase:
                print(jsonname,casename)
                jsonpath = os.path.join(json_path,jsonname)
                #print(jsonpath)
                casepath=os.path.join(case_path,casename)
                #print(casepath)
                # read and parse json file

                dict_json = json_parser(jsonpath)
                dict_json_keys = dict_json.keys()
                #print(dict_json)

                # change settings in excel based on json
                wb = openpyxl.load_workbook(casepath)
                sheet = wb['Setting']

                row = 3
                while True:
                    cell_val = sheet.cell(row, 1).value
                    if cell_val is None:
                        break
                    else:
                        if cell_val in dict_json_keys:
                            sheet.cell(row, 2).value = dict_json[cell_val]
                        row = row + 1
                wb.save(casepath)
                wb.close()


if __name__ == "__main__":

    #commoncase_path="F:\\9x07_workspace\\GW\\EC20CEHBR06A05V03M1G_GW_factory\\common_20181205_1458"
    #customcase_path ="F:\\9x07_workspace\\GW\\EC20CEHBR06A05V03M1G_GW_factory\\custom_20181205_1458"
    #json_path="F:\\9x07_workspace\\GW\\EC20CEHBR06A05V03M1G_GW_factory\\EC20CEHBR06A05V03M1G_GW_20171217"
    #json_path = "\\\\192.168.11.252\quectel\研发部\测试部\软件测试\ST-4G\\4G_QAT\QAT_9x07_GW_Excel\EC20CEHBR06A05V03M1G_GW_20171217"

    commoncase_path = sys.argv[1]
    customcase_path=sys.argv[2]
    json_path = sys.argv[3]

    AllAutoSetting(commoncase_path,json_path)
    AllAutoSetting(customcase_path,json_path)
